#!/usr/bin/env python3
"""Sweep BW1100 Single-machine designs over GPU count and intra-node bandwidth.

Each case obeys TP * PP * DP * EP * CP == num_procs. From 1024 GPUs onward
TP=8, PP=4, EP=16, CP=1 are fixed; DP and global batch double together.
Default scale: 32, 64, ..., 8192 GPUs.
"""
from __future__ import annotations

import argparse
import json
import math
import re
import time
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


CASES = (
    (32, 4, 2, 1, 4, 1),
    (64, 4, 2, 1, 8, 1),
    (128, 8, 2, 1, 8, 1),
    (256, 8, 2, 1, 16, 1),
    (512, 8, 4, 1, 16, 1),
    (1024, 8, 4, 2, 16, 1),
    (2048, 8, 4, 4, 16, 1),
    (4096, 8, 4, 8, 16, 1),
    (8192, 8, 4, 16, 16, 1),
)

HEADERS = [
    "Model", "Network", "Intra-node Bandwidth (GB/s)", "datatype",
    "TP", "PP", "DP", "EP", "CP", "Batch Size", "Microbatch Size",
    "Activation recompute", "Optimizer sharding", "Batch Time(s)",
    "Comm Time(s)", "Per-batch EP communication time(s)", "Comm Ratio",
    "Memory(GiB)", "MFU", "Linear Scaling Throughput (samples/s)",
]


def payload(nproc, tp, pp, dp, ep, cp, intra_bw):
    assert tp * pp * dp * ep * cp == nproc
    return {
        "gpu": {
            "name": "BW1100",
            "sparse_tensor_fp16_processing_power": 354,
            "sparse_tensor_fp32_processing_power": 44,
            "memory": 144,
            "memory_bandwidth": 2400,
            "bus_bandwidth": intra_bw,
            "network_bandwidth": 25,
            "support_p2p": True,
            "num_procs": nproc,
        },
        "network": {
            "network_bandwidth": 25,
            "network_topology": "Single machine",
        },
        "model": {
            "name": "DeepSeek-V3 671B", "seq_size": 4096, "hidden": 7168,
            "feedforward": 18432, "attn_heads": 128, "kv_heads": None,
            "attn_size": 128, "rope_theta": None, "rms_norm": None,
            "qk_norm": None, "ffn_type": None, "untied_embeddings": None,
            "num_blocks": 61, "vocab_size": 129280, "num_experts": 256,
            "moe_topk": 8, "norm_topk_prob": None,
            "router_aux_loss_coef": None, "num_shared_experts": 1,
            "moe_feedforward": 2048, "first_k_dense": 3,
            "moe_layer_freq": 1, "kv_size": 576, "q_lora_rank": 1536,
            "kv_lora_rank": 512, "qk_nope_head_dim": 128,
            "qk_rope_head_dim": 64, "v_head_dim": 128,
        },
        "trainning_config": {
            "optimization_strategy": "Full recomputation",
            "activation_recompute": "full", "optimizer_sharding": False,
            "tensor_par": tp, "pipeline_par": pp, "data_par": dp,
            "expert_par": ep, "context_par": cp,
            "batch_size": dp, "microbatch_size": 1,
            "matrix_dtype": "float8", "vector_dtype": "float8",
        },
    }


def get_path(obj, *keys):
    for key in keys:
        if not isinstance(obj, dict):
            return None
        obj = obj.get(key)
    return obj


def number(obj, *paths):
    for path in paths:
        value = get_path(obj, *path)
        if value is None:
            continue
        if isinstance(value, str):
            match = re.search(r"[-+]?\d+(?:\.\d+)?", value)
            if match:
                return float(match.group(0))
        return value
    return None



def row_from_result(result, nproc, tp, pp, dp, ep, cp, intra_bw):
    batch_time = number(result, ("summary", "batch_total_time"))
    comm_time = number(result, ("communication", "total_comm_time"))
    return {
        "Model": "DeepSeek-V3 671B",
        "Network": f"Single machine ({nproc} DCUs)",
        "Intra-node Bandwidth (GB/s)": intra_bw,
        "datatype": "FP8 matrix / FP8 vector",
        "TP": tp, "PP": pp, "DP": dp, "EP": ep, "CP": cp,
        "Batch Size": dp, "Microbatch Size": 1,
        "Activation recompute": "full", "Optimizer sharding": False,
        "Batch Time(s)": batch_time, "Comm Time(s)": comm_time,
        "Per-batch EP communication time(s)": number(result, ("communication", "batch_ep_comm_time")),
        "Comm Ratio": comm_time / batch_time if batch_time not in (None, 0) and comm_time is not None else None,
        "Memory(GiB)": number(result, ("memory_usage", "overall_usage"), ("memory_usage", "overall_usage_gib")),
        "MFU": number(result, ("summary", "total_efficiency")),
        "Linear Scaling Throughput (samples/s)": number(result, ("summary", "linear_scaling_throughput")),
    }


def post_json(url, body, timeout):
    request = urllib.request.Request(url, data=json.dumps(body).encode("utf-8"),
        method="POST", headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return json.load(response)


def parse_bandwidths(args):
    if args.intra_bandwidths:
        values = [float(item.strip()) for item in args.intra_bandwidths.split(",")]
    else:
        if args.intra_bw_step <= 0:
            raise ValueError("--intra-bw-step must be positive")
        if args.intra_bw_stop < args.intra_bw_start:
            raise ValueError("--intra-bw-stop must be >= --intra-bw-start")
        count = int(math.floor((args.intra_bw_stop - args.intra_bw_start) /
            args.intra_bw_step + 1e-9))
        values = [round(args.intra_bw_start + index * args.intra_bw_step, 10)
                  for index in range(count + 1)]
    if not values or any(value <= 0 for value in values):
        raise ValueError("all intra-node bandwidths must be positive")
    return values


def write_workbook(path, records, assumptions):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"
    sheet.append(HEADERS)
    for record in records:
        sheet.append([record.get(header) for header in HEADERS])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(str(cell.value)) if cell.value is not None else 0
                        for cell in column) + 2, 42)
        sheet.column_dimensions[column[0].column_letter].width = max(width, 12)
    index = {header: position + 1 for position, header in enumerate(HEADERS)}
    for header in (
        "Intra-node Bandwidth (GB/s)", "Batch Time(s)", "Comm Time(s)",
        "Per-batch EP communication time(s)", "Comm Ratio", "Memory(GiB)",
        "MFU", "Linear Scaling Throughput (samples/s)"):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=index[header]).number_format = "0.000000"

    notes = workbook.create_sheet("assumptions")
    notes.append(["Field", "Value"])
    for cell in notes[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    for key, value in assumptions.items():
        notes.append([key, json.dumps(value, ensure_ascii=False)
                      if isinstance(value, (dict, list)) else value])
    notes.column_dimensions["A"].width = 34
    notes.column_dimensions["B"].width = 110
    notes.freeze_panes = "A2"
    workbook.save(path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--url", default=(
        "http://127.0.0.1:8000/llm_training_calculator/calculator/calculate"))
    parser.add_argument("--output-dir", type=Path,
        default=Path("test/bw1100/single_machine_intra_bw_design"))
    parser.add_argument("--max-procs", type=int, default=8192)
    parser.add_argument("--intra-bw-start", type=float, default=100.0)
    parser.add_argument("--intra-bw-stop", type=float, default=250.0)
    parser.add_argument("--intra-bw-step", type=float, default=25.0)
    parser.add_argument("--intra-bandwidths",
        help="comma-separated GB/s list; overrides start/stop/step")
    parser.add_argument("--timeout", type=float, default=180.0)
    parser.add_argument("--pause-s", type=float, default=0.2)
    parser.add_argument("--save-raw", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    cases = tuple(case for case in CASES if case[0] <= args.max_procs)
    if not cases or cases[-1][0] != args.max_procs:
        supported = ", ".join(str(case[0]) for case in CASES)
        raise ValueError(f"--max-procs must be one of: {supported}")
    bandwidths = parse_bandwidths(args)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    records, raw = [], []
    total = len(cases) * len(bandwidths)
    for intra_bw in bandwidths:
        for nproc, tp, pp, dp, ep, cp in cases:
            body = payload(nproc, tp, pp, dp, ep, cp, intra_bw)
            print(f"case {len(records) + 1}/{total}: intra={intra_bw:g} GB/s "
                  f"n={nproc} TP={tp} PP={pp} DP={dp} EP={ep} CP={cp}", flush=True)
            result = {}
            if not args.dry_run:
                try:
                    result = post_json(args.url, body, args.timeout)
                except urllib.error.HTTPError as exc:
                    detail = exc.read().decode("utf-8", "replace")
                    raise RuntimeError(
                        f"API HTTP {exc.code}; intra={intra_bw:g}, n={nproc}: {detail}") from exc
            records.append(row_from_result(
                result, nproc, tp, pp, dp, ep, cp, intra_bw))
            if args.save_raw and not args.dry_run:
                raw.append({"request": body, "response": result})
            time.sleep(args.pause_s)

    assumptions = {
        "endpoint": args.url,
        "model": "DeepSeek-V3 671B",
        "network_topology": "Single machine",
        "selection_rule": "TP * PP * DP * EP * CP == num_procs",
        "batch_rule": "microbatch_size=1; batch_size=DP",
        "scale_strategy": ("32-512 changes scale-up dimensions; 1024-8192 "
            "keeps TP=8, PP=4, EP=16, CP=1 and doubles DP/batch"),
        "intra_bandwidths_GBps": bandwidths,
        "cases": [dict(num_procs=n, tp=t, pp=p, dp=d, ep=e, cp=c)
                  for n, t, p, d, e, c in cases],
    }
    xlsx = args.output_dir / "single_machine_intra_bw_sweep.xlsx"
    write_workbook(xlsx, records, assumptions)
    if args.save_raw:
        (args.output_dir / "raw_responses.json").write_text(
            json.dumps(raw, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    (args.output_dir / "assumptions.json").write_text(
        json.dumps(assumptions, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"xlsx": str(xlsx), "rows": len(records),
                      "bandwidths_GBps": bandwidths}, ensure_ascii=False))


if __name__ == "__main__":
    main()
